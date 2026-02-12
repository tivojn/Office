#!/usr/bin/env python3
"""Build a PREMIUM Super Bowl Excel document with openpyxl — v3 polished redesign."""

import os
import subprocess
import time
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from PIL import Image as PILImage

# ─── Paths ───────────────────────────────────────────────────────────────
BASE_DIR = "/Users/zanearcher/Downloads/vibe-coding/Office/Excel"
OUTPUT_PATH = os.path.join(BASE_DIR, "Super_Bowl_Ultimate_Guide.xlsx")

# ─── COLOR PALETTE ────────────────────────────────────────────────────────
NAVY      = "0B1D3A"
GOLD      = "C9A84C"
WHITE     = "FFFFFF"
OFF_WHITE = "F8FAFC"
LIGHT_GRAY= "F1F5F9"
MED_GRAY  = "E2E8F0"
BORDER_CLR= "CBD5E1"
DARK_TEXT  = "1E293B"
MUTED     = "64748B"
SLATE     = "334155"
RED       = "DC2626"
GREEN     = "16A34A"
EAGLES_GRN= "004C54"
CHIEFS_RED= "E31837"
PURPLE    = "6B21A8"

# Chart colors (rich multi-color palette)
CHART_COLORS = [
    "1E3A5F", "B8860B", "C0392B", "2E86C1", "27AE60",
    "8E44AD", "E67E22", "16A085", "D4AC0D", "5D6D7E",
]

# ─── Helpers ──────────────────────────────────────────────────────────────
def F(c):
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

def FT(sz=10, bold=False, color=DARK_TEXT, italic=False):
    return Font(name="Calibri", size=sz, bold=bold, color=color, italic=italic)

def BD(color=MED_GRAY):
    s = Side(style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def AL(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def widths(ws, m):
    for col, w in m.items():
        ws.column_dimensions[col].width = w

def cell(ws, r, c, val=None, bg=None, ft=None, al=None, bd=None):
    """Write a cell with optional styling."""
    cl = ws.cell(row=r, column=c, value=val)
    if bg: cl.fill = F(bg)
    if ft: cl.font = ft
    if al: cl.alignment = al
    if bd: cl.border = bd
    return cl

def banner(ws, r, c1, c2, text, bg=NAVY, fc=WHITE, sz=16, ht=42, bold=True):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cl = ws.cell(row=r, column=c1)
    cl.value = text
    cl.fill = F(bg)
    cl.font = FT(sz, bold, fc)
    cl.alignment = AL()
    ws.row_dimensions[r].height = ht

def gold_line(ws, r, ncols):
    ws.row_dimensions[r].height = 3
    for c in range(1, ncols + 1):
        ws.cell(row=r, column=c).fill = F(GOLD)

def paint_bg(ws, r1, r2, c1, c2, color=NAVY):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = F(color)

def img(ws, path, anchor, w=None, h=None):
    if not os.path.exists(path):
        return False
    try:
        im = XlImage(path)
        if w: im.width = w
        if h: im.height = h
        ws.add_image(im, anchor)
        return True
    except:
        return False

def setup_print(ws, landscape=False, fit_cols=1, fit_rows=0):
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.fitToWidth = fit_cols
    ws.page_setup.fitToHeight = fit_rows
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4


# ═══════════════════════════════════════════════════════════════════════════
#  DATA
# ═══════════════════════════════════════════════════════════════════════════

SB = [
    ("I","1967","Green Bay Packers",35,"Kansas City Chiefs",10,"Bart Starr","Los Angeles Memorial Coliseum, LA"),
    ("II","1968","Green Bay Packers",33,"Oakland Raiders",14,"Bart Starr","Orange Bowl, Miami"),
    ("III","1969","New York Jets",16,"Baltimore Colts",7,"Joe Namath","Orange Bowl, Miami"),
    ("IV","1970","Kansas City Chiefs",23,"Minnesota Vikings",7,"Len Dawson","Tulane Stadium, New Orleans"),
    ("V","1971","Baltimore Colts",16,"Dallas Cowboys",13,"Chuck Howley","Orange Bowl, Miami"),
    ("VI","1972","Dallas Cowboys",24,"Miami Dolphins",3,"Roger Staubach","Tulane Stadium, New Orleans"),
    ("VII","1973","Miami Dolphins",14,"Washington Redskins",7,"Jake Scott","LA Memorial Coliseum"),
    ("VIII","1974","Miami Dolphins",24,"Minnesota Vikings",7,"Larry Csonka","Rice Stadium, Houston"),
    ("IX","1975","Pittsburgh Steelers",16,"Minnesota Vikings",6,"Franco Harris","Tulane Stadium, New Orleans"),
    ("X","1976","Pittsburgh Steelers",21,"Dallas Cowboys",17,"Lynn Swann","Orange Bowl, Miami"),
    ("XI","1977","Oakland Raiders",32,"Minnesota Vikings",14,"Fred Biletnikoff","Rose Bowl, Pasadena"),
    ("XII","1978","Dallas Cowboys",27,"Denver Broncos",10,"Martin / White","Superdome, New Orleans"),
    ("XIII","1979","Pittsburgh Steelers",35,"Dallas Cowboys",31,"Terry Bradshaw","Orange Bowl, Miami"),
    ("XIV","1980","Pittsburgh Steelers",31,"Los Angeles Rams",19,"Terry Bradshaw","Rose Bowl, Pasadena"),
    ("XV","1981","Oakland Raiders",27,"Philadelphia Eagles",10,"Jim Plunkett","Superdome, New Orleans"),
    ("XVI","1982","San Francisco 49ers",26,"Cincinnati Bengals",21,"Joe Montana","Silverdome, Pontiac"),
    ("XVII","1983","Washington Redskins",27,"Miami Dolphins",17,"John Riggins","Rose Bowl, Pasadena"),
    ("XVIII","1984","Los Angeles Raiders",38,"Washington Redskins",9,"Marcus Allen","Tampa Stadium"),
    ("XIX","1985","San Francisco 49ers",38,"Miami Dolphins",16,"Joe Montana","Stanford Stadium"),
    ("XX","1986","Chicago Bears",46,"New England Patriots",10,"Richard Dent","Superdome, New Orleans"),
    ("XXI","1987","New York Giants",39,"Denver Broncos",20,"Phil Simms","Rose Bowl, Pasadena"),
    ("XXII","1988","Washington Redskins",42,"Denver Broncos",10,"Doug Williams","Jack Murphy Stadium, SD"),
    ("XXIII","1989","San Francisco 49ers",20,"Cincinnati Bengals",16,"Jerry Rice","Joe Robbie Stadium, Miami"),
    ("XXIV","1990","San Francisco 49ers",55,"Denver Broncos",10,"Joe Montana","Superdome, New Orleans"),
    ("XXV","1991","New York Giants",20,"Buffalo Bills",19,"Ottis Anderson","Tampa Stadium"),
    ("XXVI","1992","Washington Redskins",37,"Buffalo Bills",24,"Mark Rypien","Metrodome, Minneapolis"),
    ("XXVII","1993","Dallas Cowboys",52,"Buffalo Bills",17,"Troy Aikman","Rose Bowl, Pasadena"),
    ("XXVIII","1994","Dallas Cowboys",30,"Buffalo Bills",13,"Emmitt Smith","Georgia Dome, Atlanta"),
    ("XXIX","1995","San Francisco 49ers",49,"San Diego Chargers",26,"Steve Young","Joe Robbie Stadium, Miami"),
    ("XXX","1996","Dallas Cowboys",27,"Pittsburgh Steelers",17,"Larry Brown","Sun Devil Stadium, Tempe"),
    ("XXXI","1997","Green Bay Packers",35,"New England Patriots",21,"Desmond Howard","Superdome, New Orleans"),
    ("XXXII","1998","Denver Broncos",31,"Green Bay Packers",24,"Terrell Davis","Qualcomm Stadium, SD"),
    ("XXXIII","1999","Denver Broncos",34,"Atlanta Falcons",19,"John Elway","Pro Player Stadium, Miami"),
    ("XXXIV","2000","St. Louis Rams",23,"Tennessee Titans",16,"Kurt Warner","Georgia Dome, Atlanta"),
    ("XXXV","2001","Baltimore Ravens",34,"New York Giants",7,"Ray Lewis","Raymond James Stadium"),
    ("XXXVI","2002","New England Patriots",20,"St. Louis Rams",17,"Tom Brady","Superdome, New Orleans"),
    ("XXXVII","2003","Tampa Bay Buccaneers",48,"Oakland Raiders",21,"Dexter Jackson","Qualcomm Stadium, SD"),
    ("XXXVIII","2004","New England Patriots",32,"Carolina Panthers",29,"Tom Brady","Reliant Stadium, Houston"),
    ("XXXIX","2005","New England Patriots",24,"Philadelphia Eagles",21,"Deion Branch","Alltel Stadium, Jacksonville"),
    ("XL","2006","Pittsburgh Steelers",21,"Seattle Seahawks",10,"Hines Ward","Ford Field, Detroit"),
    ("XLI","2007","Indianapolis Colts",29,"Chicago Bears",17,"Peyton Manning","Dolphin Stadium, Miami"),
    ("XLII","2008","New York Giants",17,"New England Patriots",14,"Eli Manning","Univ. of Phoenix Stadium"),
    ("XLIII","2009","Pittsburgh Steelers",27,"Arizona Cardinals",23,"Santonio Holmes","Raymond James Stadium"),
    ("XLIV","2010","New Orleans Saints",31,"Indianapolis Colts",17,"Drew Brees","Sun Life Stadium, Miami"),
    ("XLV","2011","Green Bay Packers",31,"Pittsburgh Steelers",25,"Aaron Rodgers","Cowboys Stadium, Arlington"),
    ("XLVI","2012","New York Giants",21,"New England Patriots",17,"Eli Manning","Lucas Oil Stadium, Indy"),
    ("XLVII","2013","Baltimore Ravens",34,"San Francisco 49ers",31,"Joe Flacco","Superdome, New Orleans"),
    ("XLVIII","2014","Seattle Seahawks",43,"Denver Broncos",8,"Malcolm Smith","MetLife Stadium, NJ"),
    ("XLIX","2015","New England Patriots",28,"Seattle Seahawks",24,"Tom Brady","Univ. of Phoenix Stadium"),
    ("50","2016","Denver Broncos",24,"Carolina Panthers",10,"Von Miller","Levi's Stadium, Santa Clara"),
    ("LI","2017","New England Patriots",34,"Atlanta Falcons",28,"Tom Brady","NRG Stadium, Houston"),
    ("LII","2018","Philadelphia Eagles",41,"New England Patriots",33,"Nick Foles","U.S. Bank Stadium, Mpls"),
    ("LIII","2019","New England Patriots",13,"Los Angeles Rams",3,"Julian Edelman","Mercedes-Benz Stadium"),
    ("LIV","2020","Kansas City Chiefs",31,"San Francisco 49ers",20,"Patrick Mahomes","Hard Rock Stadium, Miami"),
    ("LV","2021","Tampa Bay Buccaneers",31,"Kansas City Chiefs",9,"Tom Brady","Raymond James Stadium"),
    ("LVI","2022","Los Angeles Rams",23,"Cincinnati Bengals",20,"Cooper Kupp","SoFi Stadium, Inglewood"),
    ("LVII","2023","Kansas City Chiefs",38,"Philadelphia Eagles",35,"Patrick Mahomes","State Farm Stadium, AZ"),
    ("LVIII","2024","Kansas City Chiefs",25,"San Francisco 49ers",22,"Patrick Mahomes","Allegiant Stadium, LV"),
    ("LIX","2025","Philadelphia Eagles",40,"Kansas City Chiefs",22,"Jalen Hurts","Caesars Superdome, NO"),
]

HALFTIME = [
    ("I","1967","Univ. of Arizona & Grambling State bands","Marching Band"),
    ("II","1968","Grambling State University band","Marching Band"),
    ("III","1969","Florida A&M University band","Marching Band"),
    ("X","1976","Up with People","Performance Group"),
    ("XX","1986","Up with People","Performance Group"),
    ("XXVII","1993","Michael Jackson","Pop"),
    ("XXIX","1995","Patti LaBelle, Tony Bennett, Arturo Sandoval","Pop / Jazz"),
    ("XXX","1996","Diana Ross","Pop / R&B"),
    ("XXXIII","1999","Stevie Wonder, Gloria Estefan","Pop / R&B"),
    ("XXXIV","2000","Phil Collins, Christina Aguilera, Toni Braxton","Pop / Rock"),
    ("XXXV","2001","Aerosmith, *NSYNC, Britney Spears, Mary J. Blige","Pop / Rock"),
    ("XXXVI","2002","U2","Rock"),
    ("XXXVII","2003","Shania Twain, No Doubt, Sting","Pop / Rock"),
    ("XXXVIII","2004","Janet Jackson, Justin Timberlake, P. Diddy","Pop / R&B"),
    ("XXXIX","2005","Paul McCartney","Rock"),
    ("XL","2006","The Rolling Stones","Rock"),
    ("XLI","2007","Prince","Pop / Rock"),
    ("XLII","2008","Tom Petty and the Heartbreakers","Rock"),
    ("XLIII","2009","Bruce Springsteen & The E Street Band","Rock"),
    ("XLIV","2010","The Who","Rock"),
    ("XLV","2011","The Black Eyed Peas ft. Usher, Slash","Pop / Hip-Hop"),
    ("XLVI","2012","Madonna ft. LMFAO, Nicki Minaj, Cee Lo Green","Pop"),
    ("XLVII","2013","Beyonce ft. Destiny's Child","Pop / R&B"),
    ("XLVIII","2014","Bruno Mars ft. Red Hot Chili Peppers","Pop / Rock"),
    ("XLIX","2015","Katy Perry ft. Lenny Kravitz, Missy Elliott","Pop"),
    ("50","2016","Coldplay ft. Beyonce, Bruno Mars","Pop / Rock"),
    ("LI","2017","Lady Gaga","Pop"),
    ("LII","2018","Justin Timberlake","Pop / R&B"),
    ("LIII","2019","Maroon 5 ft. Travis Scott, Big Boi","Pop / Hip-Hop"),
    ("LIV","2020","Jennifer Lopez & Shakira","Pop / Latin"),
    ("LV","2021","The Weeknd","Pop / R&B"),
    ("LVI","2022","Dr. Dre, Snoop Dogg, Eminem, Kendrick Lamar","Hip-Hop / R&B"),
    ("LVII","2023","Rihanna","Pop / R&B"),
    ("LVIII","2024","Usher","R&B / Pop"),
    ("LIX","2025","Kendrick Lamar ft. SZA","Hip-Hop / R&B"),
]

ECON = [
    ("Total Economic Output","$1.25 Billion","Impact on New Orleans metro"),
    ("Consumer Spending","$18.6 Billion","Food, merch, parties, travel"),
    ("30-Second Ad Cost","$8.0 Million","New record — up from $7M"),
    ("Fox Ad Revenue","$600+ Million","Total telecast ad revenue"),
    ("Worker Earnings","$395 Million","Event-related industries"),
    ("Jobs Created","9,787","Temp and permanent positions"),
    ("US TV Viewership","127.7 Million","New all-time SB record"),
    ("Peak Viewership","137.7 Million","8:00-8:15 PM ET, Q2"),
    ("Halftime Viewership","133.5 Million","Kendrick Lamar performance"),
    ("Hotel Occupancy","95%+","New Orleans metro during SB week"),
    ("Avg Ticket Price","$9,800+","Secondary market average"),
    ("Cheapest Ticket","$4,500+","Upper deck / standing room"),
    ("Host City Revenue","$500+ Million","Direct spend in New Orleans"),
    ("Wings Consumed","1.45 Billion","On Super Bowl Sunday"),
    ("Pizzas Ordered","12.5 Million","Deliveries on game day"),
    ("Avocados Consumed","120+ Million lbs","For guacamole nationwide"),
]

def team_wins():
    w = {}
    for r in SB:
        w[r[2]] = w.get(r[2], 0) + 1
    return sorted(w.items(), key=lambda x: -x[1])

def decade_avgs():
    d = {}
    for r in SB:
        dec = f"{(int(r[1])//10)*10}s"
        d.setdefault(dec, []).append(r[3]+r[5])
    return {k: round(sum(v)/len(v),1) for k,v in sorted(d.items())}

def margins():
    return [(r[0],r[1],r[2],r[3],r[4],r[5],abs(r[3]-r[5])) for r in SB]


# ═══════════════════════════════════════════════════════════════════════════
#  BUILD v3
# ═══════════════════════════════════════════════════════════════════════════
print("Building PREMIUM Super Bowl workbook v3...")
wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────
# SHEET 1: COVER
# ─────────────────────────────────────────────────────────────────────────
print("  [1/6] Cover page...")
ws = wb.active
ws.title = "SUPER BOWL"
ws.sheet_properties.tabColor = NAVY
ws.sheet_view.showGridLines = False

NC = 5
widths(ws, {'A': 3, 'B': 18, 'C': 18, 'D': 18, 'E': 18, 'F': 3})

# Paint dark background (only data columns, not too many rows)
for r in range(1, 42):
    for c in range(1, NC + 2):  # A-F
        ws.cell(row=r, column=c).fill = F(NAVY)
    ws.row_dimensions[r].height = 14

# Row 1: Gold accent line
gold_line(ws, 1, NC + 1)

# Rows 2-4: spacer
for r in range(2, 5):
    ws.row_dimensions[r].height = 10

# Row 5-6: "SUPER BOWL" title
ws.merge_cells('B5:E6')
cl = ws['B5']
cl.value = "SUPER BOWL"
cl.fill = F(NAVY)
cl.font = FT(44, True, WHITE)
cl.alignment = AL()
ws.row_dimensions[5].height = 40
ws.row_dimensions[6].height = 40

# Row 7: subtitle
ws.merge_cells('B7:E7')
cl = ws['B7']
cl.value = "THE ULTIMATE GUIDE  |  1967 - 2025"
cl.fill = F(NAVY)
cl.font = FT(13, False, GOLD)
cl.alignment = AL()
ws.row_dimensions[7].height = 26

# Row 8: Gold divider
gold_line(ws, 8, NC + 1)

# Row 9: spacer
ws.row_dimensions[9].height = 12

# Rows 10-11: 3 KPI panels across B, C, D, E
kpis = [("59", "GAMES PLAYED"), ("23", "UNIQUE CHAMPIONS"), ("127.7M", "PEAK TV VIEWERS")]
kpi_cols = [(2, 2), (3, 3), (4, 5)]  # (start_col, end_col)

ws.row_dimensions[10].height = 44
ws.row_dimensions[11].height = 22

for (val, label), (sc, ec) in zip(kpis, kpi_cols):
    if sc != ec:
        ws.merge_cells(start_row=10, start_column=sc, end_row=10, end_column=ec)
        ws.merge_cells(start_row=11, start_column=sc, end_row=11, end_column=ec)
    # Value
    cl = ws.cell(row=10, column=sc)
    cl.value = val
    cl.fill = F(SLATE)
    cl.font = FT(28, True, GOLD)
    cl.alignment = AL()
    cl.border = Border(left=Side(style="thin", color=GOLD), right=Side(style="thin", color=GOLD),
                       top=Side(style="medium", color=GOLD))
    # Label
    cl = ws.cell(row=11, column=sc)
    cl.value = label
    cl.fill = F(SLATE)
    cl.font = FT(8, True, WHITE)
    cl.alignment = AL()
    cl.border = Border(left=Side(style="thin", color=GOLD), right=Side(style="thin", color=GOLD),
                       bottom=Side(style="medium", color=GOLD))

# Row 12: spacer
ws.row_dimensions[12].height = 14

# Row 13: Gold divider
gold_line(ws, 13, NC + 1)

# Row 14: spacer
ws.row_dimensions[14].height = 10

# Row 15: "SUPER BOWL LIX CHAMPION"
ws.merge_cells('B15:E15')
cl = ws['B15']
cl.value = "SUPER BOWL LIX CHAMPION"
cl.fill = F(NAVY)
cl.font = FT(10, True, MUTED)
cl.alignment = AL()
ws.row_dimensions[15].height = 20

# Row 16: Score line
ws.merge_cells('B16:E16')
cl = ws['B16']
cl.value = "PHILADELPHIA EAGLES  40 - 22  KANSAS CITY CHIEFS"
cl.fill = F(NAVY)
cl.font = FT(18, True, EAGLES_GRN)
cl.alignment = AL()
ws.row_dimensions[16].height = 36

# Row 17: MVP
ws.merge_cells('B17:E17')
cl = ws['B17']
cl.value = "MVP: Jalen Hurts  |  Caesars Superdome  |  Feb 9, 2025"
cl.fill = F(NAVY)
cl.font = FT(9, False, GOLD)
cl.alignment = AL()
ws.row_dimensions[17].height = 20

# Row 18: spacer
ws.row_dimensions[18].height = 10

# Row 19: Gold divider
gold_line(ws, 19, NC + 1)

# Row 20: spacer
ws.row_dimensions[20].height = 8

# Row 21: Contents
ws.merge_cells('B21:E21')
cl = ws['B21']
cl.value = "History  |  Statistics  |  Super Bowl LIX  |  Economic Impact  |  Halftime Shows"
cl.fill = F(NAVY)
cl.font = FT(8, False, MUTED)
cl.alignment = AL()
ws.row_dimensions[21].height = 16

# Row 22: spacer
ws.row_dimensions[22].height = 6

# Trophy image
img(ws, os.path.join(BASE_DIR, "superbowl_trophy.png"), "B23", w=300, h=225)

setup_print(ws, landscape=False, fit_cols=1)


# ─────────────────────────────────────────────────────────────────────────
# SHEET 2: HISTORY
# ─────────────────────────────────────────────────────────────────────────
print("  [2/6] History...")
ws2 = wb.create_sheet("History")
ws2.sheet_properties.tabColor = GOLD
ws2.sheet_view.showGridLines = False

# Landscape — fits 9 columns comfortably
widths(ws2, {'A': 7, 'B': 6, 'C': 20, 'D': 7, 'E': 20, 'F': 7, 'G': 7, 'H': 18, 'I': 28})
HCOLS = 9

# Title
banner(ws2, 1, 1, HCOLS, "SUPER BOWL HISTORY  |  COMPLETE RESULTS 1967-2025", NAVY, WHITE, 14, 38)
banner(ws2, 2, 1, HCOLS, "59 Championship Games  |  Every Winner, Score, MVP & Location", GOLD, NAVY, 9, 24, True)

ws2.row_dimensions[3].height = 4

# Headers
hdrs = ["SB #", "Year", "Winner", "W Score", "Loser", "L Score", "Margin", "MVP", "Location"]
for i, h in enumerate(hdrs, 1):
    cl = cell(ws2, 4, i, h, NAVY, FT(9, True, GOLD), AL(), Border(bottom=Side(style="medium", color=GOLD)))
ws2.row_dimensions[4].height = 28

# Data
for idx, row in enumerate(SB):
    r = 5 + idx
    vals = [row[0], int(row[1]), row[2], row[3], row[4], row[5], row[3]-row[5], row[6], row[7]]
    bg = LIGHT_GRAY if idx % 2 == 1 else WHITE

    for ci, v in enumerate(vals, 1):
        a = AL("center") if ci in (1,2,4,6,7) else AL("left", "center")
        cell(ws2, r, ci, v, bg, FT(9, False, DARK_TEXT), a, BD(MED_GRAY))
    ws2.row_dimensions[r].height = 19

# Highlight LIX row
last = 5 + len(SB) - 1
for ci in range(1, HCOLS + 1):
    cl = ws2.cell(row=last, column=ci)
    cl.fill = F(EAGLES_GRN)
    cl.font = FT(9, True, WHITE)

ws2.freeze_panes = "A5"
setup_print(ws2, landscape=True, fit_cols=1)


# ─────────────────────────────────────────────────────────────────────────
# SHEET 3: STATISTICS
# ─────────────────────────────────────────────────────────────────────────
print("  [3/6] Statistics...")
ws3 = wb.create_sheet("Statistics")
ws3.sheet_properties.tabColor = "3B82F6"
ws3.sheet_view.showGridLines = False

widths(ws3, {'A': 24, 'B': 8, 'C': 3, 'D': 24, 'E': 8, 'F': 3, 'G': 24, 'H': 8})

banner(ws3, 1, 1, 8, "SUPER BOWL STATISTICS & ANALYSIS", NAVY, WHITE, 14, 38)
banner(ws3, 2, 1, 8, "Team Championships  |  Scoring Trends  |  Blowouts vs Closest Games", GOLD, NAVY, 9, 24, True)

ws3.row_dimensions[3].height = 4

# ── LEFT: Team Wins Table ──
ws3.merge_cells('A4:B4')
cell(ws3, 4, 1, "CHAMPIONSHIPS BY TEAM", SLATE, FT(10, True, WHITE), AL("left"))
ws3.cell(row=4, column=2).fill = F(SLATE)
ws3.row_dimensions[4].height = 26

cell(ws3, 5, 1, "Team", NAVY, FT(9, True, GOLD), AL("left"))
cell(ws3, 5, 2, "Wins", NAVY, FT(9, True, GOLD), AL("center"))
ws3.row_dimensions[5].height = 24

tw = team_wins()
for idx, (team, wins) in enumerate(tw):
    r = 6 + idx
    bg = LIGHT_GRAY if idx % 2 == 1 else WHITE
    cell(ws3, r, 1, team, bg, FT(9), AL("left", "center"), BD(MED_GRAY))
    cell(ws3, r, 2, wins, bg, FT(9, True, NAVY), AL("center"), BD(MED_GRAY))
    ws3.row_dimensions[r].height = 18

# ── RIGHT: Bar Chart (multi-color) ──
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Super Bowl Championships by Team"
chart1.y_axis.title = None
chart1.x_axis.title = None
chart1.style = 10
chart1.width = 18
chart1.height = 13
chart1.legend = None
chart1.y_axis.majorGridlines = None

top_n = min(10, len(tw))
top_end = 5 + top_n
d = Reference(ws3, min_col=2, min_row=5, max_row=top_end)
cats = Reference(ws3, min_col=1, min_row=6, max_row=top_end)
chart1.add_data(d, titles_from_data=True)
chart1.set_categories(cats)

# Color each bar differently
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
for i in range(top_n):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = CHART_COLORS[i % len(CHART_COLORS)]
    chart1.series[0].data_points.append(pt)

ws3.add_chart(chart1, "D4")

# ── LEFT (below team wins): Decade Averages ──
da = decade_avgs()
da_start = 6 + len(tw) + 2

ws3.merge_cells(f'A{da_start}:B{da_start}')
cell(ws3, da_start, 1, "AVG TOTAL SCORE BY DECADE", SLATE, FT(10, True, WHITE), AL("left"))
ws3.cell(row=da_start, column=2).fill = F(SLATE)
ws3.row_dimensions[da_start].height = 26

cell(ws3, da_start+1, 1, "Decade", NAVY, FT(9, True, GOLD), AL("left"))
cell(ws3, da_start+1, 2, "Avg Pts", NAVY, FT(9, True, GOLD), AL("center"))
ws3.row_dimensions[da_start+1].height = 24

for idx, (dec, avg) in enumerate(da.items()):
    r = da_start + 2 + idx
    bg = LIGHT_GRAY if idx % 2 == 1 else WHITE
    cell(ws3, r, 1, dec, bg, FT(9), AL("left", "center"), BD(MED_GRAY))
    cell(ws3, r, 2, avg, bg, FT(9, True, NAVY), AL("center"), BD(MED_GRAY))
    ws3.row_dimensions[r].height = 18

# Line chart for decades
chart2 = LineChart()
chart2.title = "Average Combined Score by Decade"
chart2.y_axis.title = None
chart2.width = 18
chart2.height = 11
chart2.style = 10
chart2.legend = None

dec_d = Reference(ws3, min_col=2, min_row=da_start+1, max_row=da_start+1+len(da))
dec_c = Reference(ws3, min_col=1, min_row=da_start+2, max_row=da_start+1+len(da))
chart2.add_data(dec_d, titles_from_data=True)
chart2.set_categories(dec_c)
chart2.series[0].graphicalProperties.line.solidFill = GOLD
chart2.series[0].graphicalProperties.line.width = 28000
chart2.series[0].smooth = True

# Add markers
chart2.series[0].marker.symbol = "circle"
chart2.series[0].marker.size = 7
chart2.series[0].marker.graphicalProperties.solidFill = NAVY

ws3.add_chart(chart2, f"D{da_start}")

# ── Blowouts & Closest Games (below decade data) ──
m = margins()
blowouts = sorted(m, key=lambda x: -x[6])[:8]
closest = sorted(m, key=lambda x: x[6])[:8]

blow_start = da_start + 2 + len(da) + 2

# Blowouts (left side)
ws3.merge_cells(f'A{blow_start}:B{blow_start}')
cell(ws3, blow_start, 1, "BIGGEST BLOWOUTS", RED, FT(10, True, WHITE), AL("left"))
ws3.cell(row=blow_start, column=2).fill = F(RED)
ws3.row_dimensions[blow_start].height = 26

cell(ws3, blow_start+1, 1, "Game", NAVY, FT(8, True, GOLD), AL("left"))
cell(ws3, blow_start+1, 2, "Score", NAVY, FT(8, True, GOLD), AL("center"))
ws3.row_dimensions[blow_start+1].height = 22

for idx, b in enumerate(blowouts):
    r = blow_start + 2 + idx
    bg = LIGHT_GRAY if idx % 2 == 1 else WHITE
    cell(ws3, r, 1, f"SB {b[0]}: {b[2]}", bg, FT(8), AL("left", "center"), BD(MED_GRAY))
    cell(ws3, r, 2, f"{b[3]}-{b[5]}", bg, FT(8, True, RED), AL("center"), BD(MED_GRAY))
    ws3.row_dimensions[r].height = 17

# Closest Games (right side D-E)
ws3.merge_cells(f'D{blow_start}:E{blow_start}')
cell(ws3, blow_start, 4, "CLOSEST GAMES", GREEN, FT(10, True, WHITE), AL("left"))
ws3.cell(row=blow_start, column=5).fill = F(GREEN)

cell(ws3, blow_start+1, 4, "Game", NAVY, FT(8, True, GOLD), AL("left"))
cell(ws3, blow_start+1, 5, "Score", NAVY, FT(8, True, GOLD), AL("center"))

for idx, c in enumerate(closest):
    r = blow_start + 2 + idx
    bg = LIGHT_GRAY if idx % 2 == 1 else WHITE
    cell(ws3, r, 4, f"SB {c[0]}: {c[2]}", bg, FT(8), AL("left", "center"), BD(MED_GRAY))
    cell(ws3, r, 5, f"{c[3]}-{c[5]}", bg, FT(8, True, GREEN), AL("center"), BD(MED_GRAY))

setup_print(ws3, landscape=True, fit_cols=1)


# ─────────────────────────────────────────────────────────────────────────
# SHEET 4: SUPER BOWL LIX
# ─────────────────────────────────────────────────────────────────────────
print("  [4/6] Super Bowl LIX...")
ws4 = wb.create_sheet("Super Bowl LIX")
ws4.sheet_properties.tabColor = EAGLES_GRN
ws4.sheet_view.showGridLines = False

NC4 = 6
widths(ws4, {'A': 2, 'B': 16, 'C': 16, 'D': 8, 'E': 16, 'F': 16, 'G': 2})

# Dark bg
paint_bg(ws4, 1, 45, 1, NC4 + 1, NAVY)
for r in range(1, 46):
    ws4.row_dimensions[r].height = 14

# Row 1: gold line
gold_line(ws4, 1, NC4 + 1)

# Row 2: Title
ws4.merge_cells('A2:G2')
cl = ws4['A2']
cl.value = "SUPER BOWL LIX  |  FEBRUARY 9, 2025"
cl.fill = F(NAVY)
cl.font = FT(20, True, WHITE)
cl.alignment = AL()
ws4.row_dimensions[2].height = 44

# Row 3: Subtitle
ws4.merge_cells('A3:G3')
cl = ws4['A3']
cl.value = "Caesars Superdome  |  New Orleans, Louisiana"
cl.fill = F(NAVY)
cl.font = FT(10, False, GOLD)
cl.alignment = AL()
ws4.row_dimensions[3].height = 22

# Row 4: gold line
gold_line(ws4, 4, NC4 + 1)

# Row 5: spacer
ws4.row_dimensions[5].height = 10

# Row 6: Team names
ws4.merge_cells('B6:C6')
cl = ws4['B6']
cl.value = "PHILADELPHIA EAGLES"
cl.fill = F(EAGLES_GRN)
cl.font = FT(14, True, WHITE)
cl.alignment = AL()

cell(ws4, 6, 4, "VS", NAVY, FT(10, True, GOLD), AL())

ws4.merge_cells('E6:F6')
cl = ws4['E6']
cl.value = "KANSAS CITY CHIEFS"
cl.fill = F(CHIEFS_RED)
cl.font = FT(14, True, WHITE)
cl.alignment = AL()
ws4.row_dimensions[6].height = 34

# Row 7: Scores
ws4.merge_cells('B7:C7')
cl = ws4['B7']
cl.value = 40
cl.fill = F(EAGLES_GRN)
cl.font = FT(40, True, GOLD)
cl.alignment = AL()

cell(ws4, 7, 4, "\u2014", NAVY, FT(24, True, MUTED), AL())

ws4.merge_cells('E7:F7')
cl = ws4['E7']
cl.value = 22
cl.fill = F(CHIEFS_RED)
cl.font = FT(40, True, WHITE)
cl.alignment = AL()
ws4.row_dimensions[7].height = 58

# Row 8: gold line
gold_line(ws4, 8, NC4 + 1)

# Row 9: MVP bar
ws4.merge_cells('A9:G9')
cl = ws4['A9']
cl.value = "MVP: JALEN HURTS  |  3 TDs (2 Pass, 1 Rush)  |  72 Rush Yds (QB Record)"
cl.fill = F(SLATE)
cl.font = FT(10, True, GOLD)
cl.alignment = AL()
ws4.row_dimensions[9].height = 28

# Row 10: spacer
ws4.row_dimensions[10].height = 10

# Row 11: Highlights header
ws4.merge_cells('B11:F11')
cl = ws4['B11']
cl.value = "GAME HIGHLIGHTS"
cl.fill = F(SLATE)
cl.font = FT(11, True, WHITE)
cl.alignment = AL("left")
ws4.row_dimensions[11].height = 26

highlights = [
    "Eagles scored the first 34 points, building a dominant 34-0 lead",
    "Jalen Hurts set the Super Bowl record for rushing yards by a QB (72 yds)",
    "Saquon Barkley dominated with 167 rushing yards and 2 touchdowns",
    "Philadelphia's defense shut down Patrick Mahomes in the first half",
    "Eagles denied the Chiefs an unprecedented Super Bowl three-peat",
    "Travis Kelce held to just 39 receiving yards by Eagles secondary",
    "Kendrick Lamar & SZA delivered a spectacular halftime performance",
    "127.7 million US viewers — a new Super Bowl viewership record",
]

for i, h in enumerate(highlights):
    r = 12 + i
    ws4.merge_cells(f'B{r}:F{r}')
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    cl = ws4[f'B{r}']
    cl.value = f"  \u25B8  {h}"
    cl.fill = F(bg)
    cl.font = FT(9, False, DARK_TEXT)
    cl.alignment = AL("left", "center")
    ws4.row_dimensions[r].height = 22

# Scoring by quarter
qr = 12 + len(highlights) + 1
ws4.row_dimensions[qr - 1].height = 10

ws4.merge_cells(f'B{qr}:F{qr}')
cl = ws4[f'B{qr}']
cl.value = "SCORING BY QUARTER"
cl.fill = F(SLATE)
cl.font = FT(11, True, WHITE)
cl.alignment = AL("left")
ws4.row_dimensions[qr].height = 26

# Quarter headers
q_hdrs = ["Team", "Q1", "Q2", "Q3", "Q4", "FINAL"]
for i, h in enumerate(q_hdrs):
    cell(ws4, qr+1, 2+i, h, NAVY, FT(9, True, GOLD), AL())
ws4.row_dimensions[qr+1].height = 24

# Eagles row
for i, v in enumerate(["EAGLES", 7, 20, 7, 6, 40]):
    cell(ws4, qr+2, 2+i, v, EAGLES_GRN, FT(11, True, WHITE), AL())
ws4.row_dimensions[qr+2].height = 28

# Chiefs row
for i, v in enumerate(["CHIEFS", 0, 0, 6, 16, 22]):
    cell(ws4, qr+3, 2+i, v, CHIEFS_RED, FT(11, True, WHITE), AL())
ws4.row_dimensions[qr+3].height = 28

# Images
img_r = qr + 5
ws4.row_dimensions[qr + 4].height = 10
img(ws4, os.path.join(BASE_DIR, "superbowl_action.png"), f"B{img_r}", w=340, h=220)
img(ws4, os.path.join(BASE_DIR, "superbowl_stadium.png"), f"B{img_r + 14}", w=340, h=220)

setup_print(ws4, landscape=False, fit_cols=1)


# ─────────────────────────────────────────────────────────────────────────
# SHEET 5: ECONOMIC IMPACT
# ─────────────────────────────────────────────────────────────────────────
print("  [5/6] Economic Impact...")
ws5 = wb.create_sheet("Economic Impact")
ws5.sheet_properties.tabColor = GOLD
ws5.sheet_view.showGridLines = False

widths(ws5, {'A': 2, 'B': 26, 'C': 16, 'D': 30, 'E': 2})
NC5 = 5

# Paint bg on borders only
for r in range(1, 48):
    ws5.cell(row=r, column=1).fill = F(NAVY)
    ws5.cell(row=r, column=NC5).fill = F(NAVY)
    ws5.row_dimensions[r].height = 14

# Row 1: gold line
gold_line(ws5, 1, NC5)

# Title
ws5.merge_cells('A2:E2')
cl = ws5['A2']
cl.value = "SUPER BOWL ECONOMIC IMPACT"
cl.fill = F(NAVY)
cl.font = FT(18, True, WHITE)
cl.alignment = AL()
ws5.row_dimensions[2].height = 42

ws5.merge_cells('A3:E3')
cl = ws5['A3']
cl.value = "The Business of America's Biggest Sporting Event  |  SB LIX (2025)"
cl.fill = F(NAVY)
cl.font = FT(9, False, GOLD)
cl.alignment = AL()
ws5.row_dimensions[3].height = 22

# Row 4: gold line
gold_line(ws5, 4, NC5)

# KPI panels — 2x2 grid
kpi_data = [
    ("$1.25B", "TOTAL ECONOMIC OUTPUT"),
    ("$18.6B", "CONSUMER SPENDING"),
    ("127.7M", "RECORD TV VIEWERS"),
    ("$8.0M", "30-SEC AD COST"),
]

ws5.row_dimensions[5].height = 8
row = 6
for i, (val, label) in enumerate(kpi_data):
    # Value row
    ws5.merge_cells(f'B{row}:D{row}')
    cl = ws5[f'B{row}']
    cl.value = val
    cl.fill = F(NAVY)
    cl.font = FT(26, True, GOLD)
    cl.alignment = AL()
    cl.border = Border(top=Side(style="medium", color=GOLD),
                       left=Side(style="thin", color=SLATE),
                       right=Side(style="thin", color=SLATE))
    ws5.row_dimensions[row].height = 42

    # Label row
    ws5.merge_cells(f'B{row+1}:D{row+1}')
    cl = ws5[f'B{row+1}']
    cl.value = label
    cl.fill = F(NAVY)
    cl.font = FT(9, True, WHITE)
    cl.alignment = AL()
    cl.border = Border(bottom=Side(style="medium", color=GOLD),
                       left=Side(style="thin", color=SLATE),
                       right=Side(style="thin", color=SLATE))
    ws5.row_dimensions[row+1].height = 22

    # Spacer
    ws5.row_dimensions[row+2].height = 6
    row += 3

# Full data table
row += 1
ws5.merge_cells(f'B{row}:D{row}')
cell(ws5, row, 2, "COMPLETE ECONOMIC DATA", SLATE, FT(11, True, WHITE), AL("left"))
for c in range(3, 5):
    ws5.cell(row=row, column=c).fill = F(SLATE)
ws5.row_dimensions[row].height = 28

row += 1
for i, h in enumerate(["Category", "Value", "Details"], 2):
    cell(ws5, row, i, h, NAVY, FT(9, True, GOLD), AL("left"))
ws5.row_dimensions[row].height = 24

for idx, (cat, val, det) in enumerate(ECON):
    r = row + 1 + idx
    bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
    cell(ws5, r, 2, cat, bg, FT(9, False, DARK_TEXT), AL("left", "center"), BD(MED_GRAY))
    cell(ws5, r, 3, val, bg, FT(9, True, NAVY), AL("center"), BD(MED_GRAY))
    cell(ws5, r, 4, det, bg, FT(8, False, MUTED), AL("left", "center"), BD(MED_GRAY))
    ws5.row_dimensions[r].height = 20

setup_print(ws5, landscape=False, fit_cols=1)


# ─────────────────────────────────────────────────────────────────────────
# SHEET 6: HALFTIME SHOWS
# ─────────────────────────────────────────────────────────────────────────
print("  [6/6] Halftime Shows...")
ws6 = wb.create_sheet("Halftime Shows")
ws6.sheet_properties.tabColor = PURPLE
ws6.sheet_view.showGridLines = False

widths(ws6, {'A': 9, 'B': 6, 'C': 44, 'D': 14})
NC6 = 4

banner(ws6, 1, 1, NC6, "SUPER BOWL HALFTIME SHOW HISTORY", NAVY, WHITE, 14, 38)
banner(ws6, 2, 1, NC6, "The World's Biggest Stage  |  Iconic Performances", GOLD, NAVY, 9, 24, True)

ws6.row_dimensions[3].height = 4

# Headers
for i, h in enumerate(["Super Bowl", "Year", "Performer(s)", "Genre"], 1):
    cell(ws6, 4, i, h, NAVY, FT(9, True, GOLD), AL("left" if i == 3 else "center"),
         Border(bottom=Side(style="medium", color=GOLD)))
ws6.row_dimensions[4].height = 26

for idx, (sb, yr, perf, genre) in enumerate(HALFTIME):
    r = 5 + idx
    bg = LIGHT_GRAY if idx % 2 == 1 else WHITE
    for ci, v in enumerate([sb, int(yr), perf, genre], 1):
        a = AL("left" if ci == 3 else "center", "center")
        cell(ws6, r, ci, v, bg, FT(9, False, DARK_TEXT), a, BD(MED_GRAY))
    ws6.row_dimensions[r].height = 19

# Highlight LIX row
last_h = 5 + len(HALFTIME) - 1
for ci in range(1, NC6 + 1):
    cl = ws6.cell(row=last_h, column=ci)
    cl.fill = F(PURPLE)
    cl.font = FT(9, True, WHITE)

# Genre breakdown
genres = {}
for _, _, _, g in HALFTIME:
    genres[g] = genres.get(g, 0) + 1
gs = sorted(genres.items(), key=lambda x: -x[1])

gr = last_h + 3
ws6.merge_cells(f'A{gr}:D{gr}')
cell(ws6, gr, 1, "HALFTIME SHOW GENRE BREAKDOWN", SLATE, FT(10, True, WHITE), AL("left"))
for c in range(2, NC6 + 1):
    ws6.cell(row=gr, column=c).fill = F(SLATE)
ws6.row_dimensions[gr].height = 26

# Genre header row with "Genre" label for pie chart reference
cell(ws6, gr+1, 1, "Genre", NAVY, FT(9, True, GOLD), AL("left"))
cell(ws6, gr+1, 2, "Count", NAVY, FT(9, True, GOLD), AL("center"))
ws6.row_dimensions[gr+1].height = 22

for idx, (genre, cnt) in enumerate(gs):
    r = gr + 2 + idx
    bg = LIGHT_GRAY if idx % 2 == 1 else WHITE
    cell(ws6, r, 1, genre, bg, FT(9), AL("left", "center"), BD(MED_GRAY))
    cell(ws6, r, 2, cnt, bg, FT(9, True, NAVY), AL("center"), BD(MED_GRAY))
    ws6.row_dimensions[r].height = 18

# Pie chart — FIXED: use categories properly so labels show genre names
pie = PieChart()
pie.title = "Halftime Show Genres"
pie.width = 14
pie.height = 10
pie.style = 10

# Data = count column, Categories = genre names
pie_data = Reference(ws6, min_col=2, min_row=gr+1, max_row=gr+1+len(gs))
pie_cats = Reference(ws6, min_col=1, min_row=gr+2, max_row=gr+1+len(gs))
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)

# Data labels: show category name + percent, NOT series name
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True
pie.dataLabels.showVal = False
pie.dataLabels.showSerName = False

# Color each slice
pie_colors = ["1E3A5F", "B8860B", "C0392B", "2E86C1", "27AE60",
              "8E44AD", "E67E22", "16A085", "D4AC0D", "5D6D7E", "E74C3C"]
for i in range(len(gs)):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = pie_colors[i % len(pie_colors)]
    pie.series[0].data_points.append(pt)

ws6.add_chart(pie, f"C{gr+1}")

# Halftime image
img_row = gr + 2 + len(gs) + 2
img(ws6, os.path.join(BASE_DIR, "superbowl_halftime.png"), f"A{img_row}", w=400, h=260)

ws6.freeze_panes = "A5"
setup_print(ws6, landscape=False, fit_cols=1)


# ═══════════════════════════════════════════════════════════════════════════
#  SAVE & OPEN
# ═══════════════════════════════════════════════════════════════════════════
print(f"\nSaving to {OUTPUT_PATH}...")
wb.save(OUTPUT_PATH)
print("Saved successfully!")

# Close existing and reopen fresh
subprocess.run(['osascript', '-e',
    'tell application "Microsoft Excel" to quit saving no'],
    capture_output=True)
time.sleep(1.5)
subprocess.run(['open', OUTPUT_PATH])
print("Opened in Excel!")
